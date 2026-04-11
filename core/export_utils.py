from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime


class ExcelExporter:
    """Classe utilitaire pour exporter des données en Excel"""
    
    def __init__(self, title="Export"):
        self.title = title
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = title
        self.current_row = 1
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def write_header(self, columns):
        """Écrit les en-têtes de colonnes"""
        for col, title in enumerate(columns, 1):
            cell = self.ws.cell(row=self.current_row, column=col, value=title)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        self.current_row += 1
    
    def write_rows(self, data, fields):
        """Écrit les lignes de données"""
        for row in data:
            for col, field in enumerate(fields, 1):
                value = self._get_nested_value(row, field)
                cell = self.ws.cell(row=self.current_row, column=col, value=value)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center')
            self.current_row += 1
    
    def _get_nested_value(self, obj, field):
        """Accède aux valeurs imbriquées (ex: 'eleve.nom')"""
        parts = field.split('.')
        value = obj
        for part in parts:
            if hasattr(value, part):
                value = getattr(value, part)
                if callable(value):
                    value = value()
            else:
                return ""
            if value is None:
                return ""
        return value
    
    def auto_size_columns(self, num_cols):
        """Ajuste automatiquement la largeur des colonnes"""
        for col in range(1, num_cols + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in self.ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column].width = adjusted_width
    
    def get_response(self, filename=None):
        """Retourne une réponse HTTP avec le fichier Excel"""
        if filename is None:
            filename = f"{self.title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        self.wb.save(response)
        return response


def export_eleves(request, eleves):
    """Exporte les élèves en Excel"""
    exporter = ExcelExporter("Élèves")
    columns = ['Matricule', 'Nom', 'Prénom', 'Sexe', 'Date de naissance', 'Classe', 'Statut']
    exporter.write_header(columns)
    
    fields = ['matricule', 'nom', 'prenom', 'sexe', 'date_naissance', 'classe.nom', 'statut']
    exporter.write_rows(eleves, fields)
    exporter.auto_size_columns(len(columns))
    return exporter.get_response("eleves_export")


def export_paiements(request, paiements):
    """Exporte les paiements en Excel"""
    exporter = ExcelExporter("Paiements")
    columns = ['Date', 'Élève', 'Montant', 'Mode paiement', 'Référence', 'Personnel']
    exporter.write_header(columns)
    
    fields = ['date_paiement', 'eleve.nom_complet', 'montant', 'mode_paiement', 'reference', 'personnel']
    exporter.write_rows(paiements, fields)
    exporter.auto_size_columns(len(columns))
    return exporter.get_response("paiements_export")


def export_personnel(request, personnel_list):
    """Exporte le personnel en Excel"""
    exporter = ExcelExporter("Personnel")
    columns = ['Nom', 'Prénom', 'Fonction', 'Téléphone', 'Salaire', 'Statut']
    exporter.write_header(columns)
    
    fields = ['nom', 'prenom', 'fonction', 'telephone', 'salaire_mensuel', 'est_actif']
    exporter.write_rows(personnel_list, fields)
    exporter.auto_size_columns(len(columns))
    return exporter.get_response("personnel_export")